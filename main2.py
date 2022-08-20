import openpyxl
from sys import argv

def main():
    # nomes dos arquivos
    arquivo_dados = argv[1]
    arquivo_separado = argv[2]

    book = openpyxl.load_workbook(arquivo_dados) # carrega a planilha com os dados
    book2 = openpyxl.Workbook(arquivo_separado) # cria uma planilha vazia

    paginas_planilha = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Erro'] # páginas da planilha

    # cria as páginas da planilha
    for pagina in paginas_planilha:
        book2.create_sheet(pagina) 
    
    # página onde estão os dados
    dados_page = book['Respostas ao formulário 1'] 
    
    marcaram_errado = []

    errado = book2['Erro']
    quant = 2

    dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']

    for dia in dias_semana:
        dia_semana = book2[dia] # selecionar a pagina da planilha pra editar
        
        # contadores
        quant += 1
        quant_mat = 0
        quant_vesp = 0

        dia_semana.append(['Nº', 'Nome', 'Matutino', 'Vespertino'])
        contador = 0 # adicionar o número na coluna 'Nº'

        for rows in dados_page.iter_rows(min_row=2, max_row=89):
            # 1 = nome, 2 = curso, 3 a 8 os dias da semana com os turnos
            if rows[quant].value == 'Matutino, Vespertino':
                if rows[1].value not in marcaram_errado:
                    marcaram_errado.append(rows[1].value)

            elif rows[quant].value == 'Matutino':
                quant_mat += 1
                contador += 1
                dia_semana.append([contador, rows[1].value, 'X', ' ']) # adiciona os dados na primeira linha

            elif rows[quant].value == 'Vespertino':
                quant_vesp += 1
                contador += 1
                dia_semana.append([contador, rows[1].value, ' ', 'X']) # adiciona os dados na primeira linha

        dia_semana.append(['Total', ' ', quant_mat, quant_vesp]) # a quantidade de pessoas em cada turno no dia
    
    # separa os nomes de quem preencheu o formulário de forma errada
    for nome in marcaram_errado:
        errado.append([nome])

    # salva a planilha
    book2.save(arquivo_separado)

    print('Dados filtrados e separados!')
    
if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(e)